from fastapi import APIRouter, Depends, HTTPException, Query, Request, UploadFile, File
from sqlalchemy import select, func, cast, String, or_, and_, Integer
from sqlalchemy.orm import selectinload
from sqlalchemy.ext.asyncio import AsyncSession
from fastapi.responses import StreamingResponse
import io
import csv
import uuid
import os
import pandas as pd
from datetime import datetime
from fastapi.responses import FileResponse
from typing import Optional, List, Union
import openpyxl

# Importaciones locales
from src.database.session import get_db
from src.database.models import Comprobante, CfdiRelacionado, CfdiConcepto
from src.api.endpoints.auth import get_active_entidad
from src.services.cfdi_storage import get_cfdi_vault_path

# --- BLINDAJE L3 ---
from src.core.license_core import VantecSystemState

# [VCORE L6] RUTAS RELATIVAS (ANTI-CRASH)
BASE_DIR = os.path.dirname(os.path.dirname(os.path.dirname(os.path.dirname(os.path.abspath(__file__)))))
LOGS_DIR = os.path.join(BASE_DIR, "Operacion_CFDI", "logs", "duplicates")
UPLOAD_DIR = os.path.join(BASE_DIR, "Operacion_CFDI", "Upload_Universal")

router = APIRouter(tags=["Comprobantes"])

@router.get("/")
@router.get("")
async def get_comprobantes(
    request: Request,
    limit: int = Query(100),
    offset: int = Query(0),
    search: str | None = Query(None),
    db: AsyncSession = Depends(get_db),
    entidad_id: uuid.UUID = Depends(get_active_entidad),
    fecha_desde: str | None = Query(None),
    fecha_hasta: str | None = Query(None),
    tipo: str | None = Query(None),
    serie: str | None = Query(None),
    folio_desde: int | None = Query(None),
    folio_hasta: int | None = Query(None)
):
    try:
        if not entidad_id:
             raise HTTPException(status_code=428, detail="CONTEXTO_RFC_OBLIGATORIO")

        query = select(Comprobante).options(selectinload(Comprobante.relacionados))

        query = query.where(Comprobante.entidad_id == entidad_id)
        if not search:
            if fecha_desde:
                try:
                    query = query.where(Comprobante.fecha_emision >= datetime.strptime(fecha_desde, '%Y-%m-%d'))
                except: pass
            if fecha_hasta:
                try:
                    query = query.where(Comprobante.fecha_emision <= datetime.strptime(fecha_hasta, '%Y-%m-%d'))
                except: pass

        if tipo:
            query = query.where(Comprobante.tipo_comprobante == tipo)

        if serie:
             query = query.where(func.trim(func.coalesce(Comprobante.serie, '')).ilike(f"%{serie.strip()}%"))
        
        if folio_desde is not None:
             clean_folio = func.regexp_replace(func.trim(Comprobante.folio), '[^0-9]', '', 'g')
             query = query.where(cast(func.nullif(clean_folio, ''), Integer) >= folio_desde)

        if folio_hasta is not None:
             clean_folio = func.regexp_replace(func.trim(Comprobante.folio), '[^0-9]', '', 'g')
             query = query.where(cast(func.nullif(clean_folio, ''), Integer) <= folio_hasta)

        if search:
            search_raw = search.strip()
            search_clean = search_raw.replace('-', '').replace(' ', '')
            search_numeric = "".join(filter(str.isdigit, search_clean)).lstrip('0')
            folio_db_clean = func.ltrim(func.trim(func.coalesce(Comprobante.folio, '')), '0')
            serie_db_clean = func.trim(func.coalesce(Comprobante.serie, ''))

            import re
            serie_folio_match = re.match(r'^([A-Za-z]+)[\-\s]*(\d+)$', search_raw)

            search_conditions = [
                func.concat(serie_db_clean, folio_db_clean).ilike(f"%{search_clean.lstrip('0')}%"),
                cast(Comprobante.uuid, String).ilike(f"%{search_raw}%"),
                func.coalesce(Comprobante.nombre_receptor, '').ilike(f"%{search_raw}%"),
                func.coalesce(Comprobante.nombre_emisor, '').ilike(f"%{search_raw}%")
            ]

            if search_numeric:
                search_conditions.append(folio_db_clean.ilike(f"%{search_numeric}%"))

            concept_subquery = (
                select(CfdiConcepto.cfdi_id)
                .where(CfdiConcepto.descripcion.ilike(f"%{search_raw}%"))
            )
            search_conditions.append(Comprobante.uuid.in_(concept_subquery))

            if serie_folio_match:
                search_serie = serie_folio_match.group(1).strip()
                search_folio = serie_folio_match.group(2).lstrip('0')
                if search_serie and search_folio:
                    search_conditions.append(
                        and_(
                            serie_db_clean.ilike(f"%{search_serie}%"),
                            folio_db_clean.ilike(f"%{search_folio}%")
                        )
                    )

            query = query.where(or_(*search_conditions))

        query = query.order_by(Comprobante.fecha_emision.desc())
        
        count_q = select(func.count()).select_from(query.subquery())
        total_records = await db.scalar(count_q)

        result = await db.execute(query.limit(limit).offset(offset))
        rows = result.scalars().all()
        
        row_uuids = [str(r.uuid) for r in rows]
        
        all_rel_uuids = []
        for r in rows:
            all_rel_uuids.extend([rel.uuid_relacionado for rel in r.relacionados])
            
        inverse_map = {}
        if row_uuids:
             inv_rels_q = await db.execute(
                  select(CfdiRelacionado, Comprobante)
                  .join(Comprobante, Comprobante.uuid == CfdiRelacionado.cfdi_id)
                  .where(func.lower(CfdiRelacionado.uuid_relacionado).in_([u.lower() for u in row_uuids]))
             )
             for row_inv in inv_rels_q.all():
                  rel_obj, rel_comp = row_inv
                  u_rel = str(rel_obj.uuid_relacionado).lower()
                  if u_rel not in inverse_map:
                       inverse_map[u_rel] = []
                  type_label = "Pago" if rel_comp.tipo_comprobante == 'P' else "Relacionado"
                  
                  p_xml = rel_comp.xml_path
                  p_pdf_raw = rel_comp.pdf_path
                  xml_ok = os.path.exists(p_xml) if p_xml else False
                  p_list = [p.replace('\\', '/').strip() for p in (p_pdf_raw or "").split('|') if p.strip()]
                  pdf_real_exists = any(os.path.exists(p) for p in p_list)
                  if not pdf_real_exists:
                       from src.services.cfdi_storage import get_cfdi_vault_path
                       try:
                           u_hijo = str(rel_comp.uuid).lower()
                           search_dirs = []
                           if rel_comp.xml_path and os.path.exists(os.path.dirname(rel_comp.xml_path)):
                               search_dirs.append(os.path.dirname(rel_comp.xml_path))
                           
                           v_path = get_cfdi_vault_path(str(rel_comp.entidad_id), str(rel_comp.rfc_emisor), rel_comp.fecha_emision)
                           search_dirs.extend([v_path, os.path.dirname(v_path), os.path.dirname(os.path.dirname(v_path))])

                           for s_dir in search_dirs:
                               if os.path.exists(s_dir):
                                   for f in os.listdir(s_dir):
                                       if f.lower().endswith('.pdf') and u_hijo in f.lower():
                                            pdf_real_exists = True
                                            p_list = [os.path.join(s_dir, f)]
                                            break
                               if pdf_real_exists: break
                       except Exception as e_deep:
                           print(f"[VCORE DEEP SCAN ERROR] Falla en asociados {rel_comp.uuid}: {str(e_deep)}")

                  inverse_map[u_rel].append({
                       "uuid": str(rel_obj.cfdi_id),
                       "monto": float(rel_obj.monto_pagado or 0),
                       "tipo_documento": type_label,
                       "folio": f"{rel_comp.serie or ''} {rel_comp.folio or 'S/N'}".strip(),
                       "rfc_receptor": rel_comp.rfc_receptor or "",
                       "pdf_exists": pdf_real_exists,
                       "pdf_count": len(p_list),
                       "pdf_files": [os.path.basename(p) for p in p_list],
                       "xml_exists": xml_ok
                  })

        rel_data_sub = {}
        if all_rel_uuids:
              rel_res = await db.execute(select(Comprobante.uuid, Comprobante.folio, Comprobante.serie, Comprobante.rfc_receptor, Comprobante.tipo_comprobante).where(func.lower(cast(Comprobante.uuid, String)).in_([u.lower() for u in all_rel_uuids])))
              for rr in rel_res.all():
                   rel_data_sub[str(rr[0]).lower()] = {"folio": rr[1], "serie": rr[2], "rfc_receptor": rr[3], "tipo": rr[4]}

        all_page_uuids = [str(r.uuid).lower() for r in rows]
        
        pagos_sum = {}
        if all_page_uuids:
             sum_p_batch = await db.execute(
                 select(cast(CfdiRelacionado.cfdi_id, String).label('uuid'), func.sum(CfdiRelacionado.monto_pagado))
                 .where(func.lower(cast(CfdiRelacionado.cfdi_id, String)).in_(all_page_uuids))
                 .group_by(cast(CfdiRelacionado.cfdi_id, String))
             )
             for row in sum_p_batch.all():
                  pagos_sum[str(row[0]).lower()] = float(row[1] or 0)
                  
        ppd_sum = {}
        if all_page_uuids:
             sum_ppd_batch = await db.execute(
                 select(func.lower(CfdiRelacionado.uuid_relacionado).label('uuid'), func.sum(CfdiRelacionado.monto_pagado))
                 .where(func.lower(CfdiRelacionado.uuid_relacionado).in_(all_page_uuids))
                 .group_by(func.lower(CfdiRelacionado.uuid_relacionado))
             )
             for row in sum_ppd_batch.all():
                  ppd_sum[str(row[0]).lower()] = float(row[1] or 0)
                  
        all_query_uuids = list(set([str(u).lower() for u in all_rel_uuids]))
        rel_paths = {}
        if all_query_uuids:
             rel_batch = await db.execute(
                 select(func.lower(cast(Comprobante.uuid, String)), Comprobante.xml_path, Comprobante.pdf_path)
                 .where(func.lower(cast(Comprobante.uuid, String)).in_(all_query_uuids))
             )
             for row in rel_batch.all():
                  rel_paths[str(row[0]).lower()] = {"xml": row[1], "pdf": row[2]}

        async_fallback_map = {}
        output = []
        for r in rows:
            try:
                uuid_lower = str(r.uuid).lower()
                total_final = float(r.total or 0)
                if r.tipo_comprobante == 'P':
                    total_final = pagos_sum.get(uuid_lower, 0.0)
    
                ruta_xml = r.xml_path or r.ruta_resguardo or ""
                ruta_pdf_raw = r.pdf_path or ""
                
                pdf_list = [p.replace('\\', '/').strip() for p in ruta_pdf_raw.split('|') if p.strip()]
                pdf_exists = any(os.path.exists(p) and os.path.isfile(p) for p in pdf_list) if pdf_list else False
                xml_exists = os.path.exists(ruta_xml) and os.path.isfile(ruta_xml) if ruta_xml else False
                
                if not xml_exists:
                    dup_xml = os.path.join(LOGS_DIR, f"{r.uuid}.xml")
                    if os.path.exists(dup_xml):
                        xml_exists = True
                        ruta_xml = dup_xml
                
                if not pdf_exists:
                    dup_pdf = os.path.join(LOGS_DIR, f"{r.uuid}.pdf")
                    if os.path.exists(dup_pdf):
                        pdf_exists = True
                        pdf_list = [dup_pdf]
                
                if not pdf_exists and r.fecha_emision:
                    try:
                        vault_dir = get_cfdi_vault_path(r.entidad_id, r.rfc_emisor, r.fecha_emision)
                        potential_pdf = os.path.join(vault_dir, f"{r.uuid}.pdf")
                        if os.path.exists(potential_pdf):
                            pdf_exists = True
                            pdf_list = [potential_pdf]
                            idx = 1
                            while True:
                                p_next = os.path.join(vault_dir, f"{r.uuid}_{idx}.pdf")
                                if os.path.exists(p_next):
                                    pdf_list.append(p_next)
                                    idx += 1
                                else:
                                    break
                    except Exception as e_vault:
                        print(f"--- ERROR VAULT RECOVERY {r.uuid} --- : {str(e_vault)}")

                if getattr(r, 'orphan_payment', False):
                    estatus_final = "AUSENTE"
                elif r.metodo_pago == 'PUE':
                    estatus_final = "Pagado"
                elif r.metodo_pago == 'PPD':
                    total_pagado = ppd_sum.get(uuid_lower, 0.0)
                    if total_pagado <= 0:
                        estatus_final = "Pendiente"
                    elif total_pagado >= float(r.total or 0):
                        estatus_final = "Pagado"
                    else:
                        estatus_final = "Parcial"
                else:
                    estatus_final = "Vigente"
    
                reps_directos = []
                for rel in r.relacionados:
                     rel_uuid = str(rel.uuid_relacionado).lower()
                     rel_tipo = rel_data_sub.get(rel_uuid, {}).get("tipo") or "I"
                     rel_folio = rel_data_sub.get(rel_uuid, {}).get("folio") or ""
                     rel_serie = rel_data_sub.get(rel_uuid, {}).get("serie") or ""
                     
                     rel_paths_info = rel_paths.get(rel_uuid, {})
                     r_xml = rel_paths_info.get("xml")
                     r_pdf = rel_paths_info.get("pdf")
                     rel_xml_exists = os.path.exists(r_xml) if r_xml else False
                     rel_pdf_exists = os.path.exists(r_pdf) if r_pdf else False
                     if not rel_xml_exists or not rel_pdf_exists:
                          f_att = async_fallback_map.get(rel_uuid, {})
                          if not rel_xml_exists and f_att.get("xml_path") and os.path.exists(f_att["xml_path"]): rel_xml_exists = True
                          if not rel_pdf_exists and f_att.get("pdf_path") and os.path.exists(f_att["pdf_path"]): rel_pdf_exists = True
                     
                     r_pdf_raw = r_pdf or ""
                     r_pdf_list = [p for p in r_pdf_raw.split('|') if p.strip()]
                     rel_pdf_exists = any(os.path.exists(p) for p in r_pdf_list)
                     
                     reps_directos.append({
                         "uuid": rel_uuid,
                         "monto": float(rel.monto_pagado or 0),
                         "tipo_documento": "Factura" if rel_tipo == 'I' else "Documento",
                         "folio": f"{rel_serie} {rel_folio}".strip() or "S/N",
                         "rfc_receptor": rel_data_sub.get(rel_uuid, {}).get("rfc_receptor", ""),
                         "pdf_exists": rel_pdf_exists,
                         "pdf_count": len(r_pdf_list),
                         "pdf_files": [os.path.basename(p) for p in r_pdf_list],
                         "xml_exists": rel_xml_exists
                     })
                
                reps_inversos = inverse_map.get(str(r.uuid).lower(), [])
                reps_all = reps_directos + reps_inversos
    
                output.append({
                    "uuid": str(r.uuid),
                    "fecha": r.fecha_emision.strftime("%d/%m/%Y") if r.fecha_emision else "---",
                    "serie": r.serie or "",
                    "folio": r.folio or "",
                    "rfc_emisor": r.rfc_emisor,
                    "nombre_emisor": r.nombre_emisor,
                    "rfc_receptor": r.rfc_receptor,
                    "nombre_receptor": r.nombre_receptor or "S/N",
                    "tipo": r.tipo_comprobante or "I",
                    "total": total_final,
                    "moneda": r.moneda or "MXN",
                    "metodo_pago": r.metodo_pago or "---",
                    "forma_pago": r.forma_pago or "---",
                    "estatus": estatus_final,
                    "tiene_relacionados": len(reps_all) > 0,
                    "reps_asociados": reps_all,
                    "xml_exists": xml_exists,
                    "pdf_exists": pdf_exists,
                    "pdf_count": len(pdf_list),
                    "pdf_files": [os.path.basename(p) for p in pdf_list],
                    "orphan_payment": r.orphan_payment
                    })
            except Exception as e_row:
                print(f"--- ERROR EN FILA {getattr(r, 'uuid', 'Desconocido')} --- : {str(e_row)}")
            
        return {
             "total_registros_bd": total_records,
             "resultados": output
        }

    except Exception as e:
        import traceback
        print(f"--- [CRITICAL] REPORT GENERATION ERROR --- \n{str(e)}")
        traceback.print_exc()
        raise HTTPException(status_code=500, detail=f"Error al listar comprobantes: {str(e)}")

@router.get("/export")
async def export_comprobantes(
    fmt: str = Query("xlsx", alias="format"),
    fecha_inicio: Optional[str] = Query(None, alias="startDate"),
    fecha_fin: Optional[str] = Query(None, alias="endDate"),
    moneda: Optional[str] = Query(None, alias="currency"),
    concepto: Optional[str] = None,
    search: Optional[str] = Query(None, alias="searchTerm"),
    db: AsyncSession = Depends(get_db),
    entidad_id: uuid.UUID = Depends(get_active_entidad)
):
    # --- CERROJO L3 VANTEC ---
    if VantecSystemState.IS_RESTRICTED:
        raise HTTPException(status_code=403, detail="🔒 MODO RESTRINGIDO: El periodo de evaluación de reportes ha finalizado. Adquiera una licencia para habilitar la descarga.")
        
    try:
        from datetime import datetime
        import pandas as pd
        import io
        from fastapi import HTTPException
        from fastapi.responses import StreamingResponse
        from sqlalchemy import func, cast, String, or_, and_, select
        from sqlalchemy.orm import selectinload
        
        query = (
            select(Comprobante)
            .options(selectinload(Comprobante.relacionados))
        )
        if entidad_id:
            query = query.where(Comprobante.entidad_id == entidad_id)
        
        if fecha_inicio and fecha_inicio.strip():
            query = query.where(Comprobante.fecha_emision >= datetime.fromisoformat(fecha_inicio))
        if fecha_fin and fecha_fin.strip():
            f_fin = fecha_fin if " " in fecha_fin else f"{fecha_fin} 23:59:59"
            query = query.where(Comprobante.fecha_emision <= datetime.fromisoformat(f_fin))
            
        if moneda and moneda != 'Todas':
            query = query.where(Comprobante.moneda == moneda)
             
        if concepto:
            query = query.join(CfdiConcepto).where(CfdiConcepto.descripcion == concepto)

        if search:
            search_raw = search.strip()
            search_clean = search_raw.replace('-', '').replace(' ', '')
            search_numeric = "".join(filter(str.isdigit, search_clean)).lstrip('0')
            folio_db_clean = func.ltrim(func.trim(func.coalesce(Comprobante.folio, '')), '0')
            serie_db_clean = func.trim(func.coalesce(Comprobante.serie, ''))
            
            search_conditions = [
                func.concat(serie_db_clean, folio_db_clean).ilike(f"%{search_clean.lstrip('0')}%"),
                cast(Comprobante.uuid, String).ilike(f"%{search_raw}%"),
                func.coalesce(Comprobante.nombre_receptor, '').ilike(f"%{search_raw}%"),
                func.coalesce(Comprobante.nombre_emisor, '').ilike(f"%{search_raw}%")
            ]
            if search_numeric:
                search_conditions.append(folio_db_clean.ilike(f"%{search_numeric}%"))
            
            concept_subquery = (
                select(CfdiConcepto.cfdi_id)
                .where(CfdiConcepto.descripcion.ilike(f"%{search_raw}%"))
            )
            search_conditions.append(Comprobante.uuid.in_(concept_subquery))
            
            query = query.where(or_(*search_conditions))

        query = query.order_by(Comprobante.fecha_emision.desc())
        result = await db.execute(query)
        rows = result.scalars().all()
        
        all_uuids = [str(r.uuid).lower() for r in rows]
        
        pagos_sum = {}
        if all_uuids:
             sum_p_batch = await db.execute(
                 select(func.lower(cast(CfdiRelacionado.cfdi_id, String)), func.sum(CfdiRelacionado.monto_pagado))
                 .where(func.lower(cast(CfdiRelacionado.cfdi_id, String)).in_(all_uuids))
                 .group_by(func.lower(cast(CfdiRelacionado.cfdi_id, String)))
             )
             for r_sum in sum_p_batch.all():
                 pagos_sum[str(r_sum[0])] = float(r_sum[1] or 0)

        ppd_sum = {}
        if all_uuids:
             sum_ppd_batch = await db.execute(
                 select(func.lower(CfdiRelacionado.uuid_relacionado), func.sum(CfdiRelacionado.monto_pagado))
                 .where(func.lower(CfdiRelacionado.uuid_relacionado).in_(all_uuids))
                 .group_by(func.lower(CfdiRelacionado.uuid_relacionado))
             )
             for r_ppd in sum_ppd_batch.all():
                 ppd_sum[str(r_ppd[0])] = float(r_ppd[1] or 0)
        
        related_info_map = {}
        if all_uuids:
             rel_lookup = await db.execute(
                 select(func.lower(CfdiRelacionado.uuid_relacionado), Comprobante.folio, Comprobante.serie)
                 .join(Comprobante, Comprobante.uuid == CfdiRelacionado.cfdi_id)
                 .where(func.lower(CfdiRelacionado.uuid_relacionado).in_(all_uuids))
             )
             for r_rel in rel_lookup.all():
                 u_rel = str(r_rel[0])
                 label = f"{r_rel[2] or ''} {r_rel[1] or ''}".strip()
                 if u_rel not in related_info_map: related_info_map[u_rel] = []
                 related_info_map[u_rel].append(label)

        data_list = []
        for r in rows:
            u_low = str(r.uuid).lower()
            
            if r.tipo_comprobante == 'P':
                estatus_vcore = "Pago Aplicado"
                monto_total_doc = pagos_sum.get(u_low, 0.0)
            elif r.metodo_pago == 'PUE':
                estatus_vcore = "PAGADA"
                monto_total_doc = float(r.total or 0)
            elif r.metodo_pago == 'PPD':
                t_pagado = ppd_sum.get(u_low, 0.0)
                if t_pagado <= 0: estatus_vcore = "PENDIENTE"
                elif t_pagado >= float(r.total or 0): estatus_vcore = "PAGADA"
                else: estatus_vcore = "PARCIAL"
                monto_total_doc = float(r.total or 0)
            else:
                estatus_vcore = "Vigente"
                monto_total_doc = float(r.total or 0)

            final_kpi_value = monto_total_doc if (r.tipo_comprobante == 'I' and r.estatus_sat == 'Vigente') else 0.0

            base_row = {
                "UUID": str(r.uuid),
                "Folio": f"{r.serie or ''}{r.folio or ''}".strip() or "S/F",
                "Fecha": r.fecha_emision.strftime("%Y-%m-%d") if r.fecha_emision else "-",
                "Emisor": r.nombre_emisor or r.rfc_emisor,
                "Receptor": r.nombre_receptor or r.rfc_receptor,
                "Tipo": r.tipo_comprobante,
                "Moneda": r.moneda or "MXN",
                "Monto Real": monto_total_doc,
                "Total Comprobante (Auditado)": final_kpi_value,
                "Subtotal": float(r.subtotal or 0),
                "Método Pago": r.metodo_pago or "-",
                "Estatus SAT": r.estatus_sat or "Vigente",
                "ESTATUS VCORE": estatus_vcore,
                "Pagos/Relaciones": ", ".join(related_info_map.get(u_low, ["-"])) if r.tipo_comprobante == 'I' else "-"
            }
            data_list.append(base_row)

        df = pd.DataFrame(data_list)
        
        if fmt == "csv":
            output = io.StringIO()
            output.write('\ufeff')
            if not df.empty:
                df.to_csv(output, index=False, encoding='utf-8')
            output.seek(0)
            filename = f"VCore_Export_{datetime.now().strftime('%Y%m%d')}.csv"
            return StreamingResponse(
                io.BytesIO(output.getvalue().encode('utf-8-sig')),
                media_type="text/csv",
                headers={"Content-Disposition": f"attachment; filename={filename}"}
            )
        
        output = io.BytesIO()
        try:
            if df.empty:
                with pd.ExcelWriter(output, engine='openpyxl') as writer:
                    pd.DataFrame(columns=["Sin Datos"]).to_excel(writer, index=False)
            else:
                with pd.ExcelWriter(output, engine='openpyxl') as writer:
                    df.to_excel(writer, index=False, sheet_name='VCore Master Audit')
                    worksheet = writer.sheets['VCore Master Audit']
                    
                    from openpyxl.styles import Font, PatternFill, Alignment
                    from openpyxl.utils import get_column_letter
 
                    blue_fill = PatternFill(start_color='1E3A5F', end_color='1E3A5F', fill_type='solid')
                    white_font = Font(color='FFFFFF', bold=True, size=11)
                    
                    for col_num, _ in enumerate(df.columns, 1):
                        cell = worksheet.cell(row=1, column=col_num)
                        cell.fill = blue_fill
                        cell.font = white_font
                        cell.alignment = Alignment(horizontal='center', vertical='center')
                    
                    try:
                        real_col_idx = list(df.columns).index("Monto Real") + 1
                        money_col_idx = list(df.columns).index("Total Comprobante (Auditado)") + 1
                        subtotal_col_idx = list(df.columns).index("Subtotal") + 1
                        
                        for r_idx in range(2, len(df) + 2):
                            worksheet.cell(row=r_idx, column=real_col_idx).number_format = '"$"#,##0.00'
                            worksheet.cell(row=r_idx, column=money_col_idx).number_format = '"$"#,##0.00'
                            worksheet.cell(row=r_idx, column=subtotal_col_idx).number_format = '"$"#,##0.00'
                        
                        total_row = len(df) + 3
                        worksheet.cell(row=total_row, column=money_col_idx - 1).value = "TOTAL AUDITADO VCORE:"
                        worksheet.cell(row=total_row, column=money_col_idx - 1).font = Font(bold=True)
                        
                        col_letter = get_column_letter(money_col_idx)
                        formula = f"=SUM({col_letter}2:{col_letter}{len(df) + 1})"
                        worksheet.cell(row=total_row, column=money_col_idx).value = formula
                        worksheet.cell(row=total_row, column=money_col_idx).font = Font(bold=True, color='27AE60', size=12)
                        worksheet.cell(row=total_row, column=money_col_idx).number_format = '"$"#,##0.00'
                    except Exception as e_inner:
                        print(f"[!] Error en columnas de Excel: {e_inner}")

                    for col in worksheet.columns:
                        max_length = 0
                        for cell in col:
                            if cell.value: max_length = max(max_length, len(str(cell.value)))
                        worksheet.column_dimensions[get_column_letter(col[0].column)].width = max_length + 2
        except Exception as e_excel:
            print(f"[CRITICAL] EXCEL_GEN_FAIL: {e_excel}")
            raise HTTPException(status_code=500, detail=f"Error interno generating Excel: {str(e_excel)}")

        output.seek(0)
        filename = f"VCore_Master_Audit_{datetime.now().strftime('%Y%m%d')}.xlsx"
        return StreamingResponse(
            output,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f"attachment; filename={filename}"}
        )

    except HTTPException:
        raise
    except Exception as e:
        import traceback
        traceback.print_exc()
        raise HTTPException(status_code=500, detail=f"Sincronización Fallida: {str(e)}")

@router.get("/{uuid}/xml")
async def get_comprobante_xml(uuid: str, db: AsyncSession = Depends(get_db)):
    if not uuid or len(uuid) < 32:
        raise HTTPException(status_code=400, detail="UUID inválido o malformado")
        
    try:
        query = select(Comprobante).where(func.lower(cast(Comprobante.uuid, String)) == uuid.lower())
        res = await db.execute(query)
        comp = res.scalars().first()
    except Exception as e_db:
        raise HTTPException(status_code=500, detail=f"Error de base de datos: {str(e_db)}")
    
    if not comp:
        raise HTTPException(status_code=404, detail="Comprobante no encontrado")

    import os
    from fastapi.responses import FileResponse
    from src.services.cfdi_storage import find_cfdi_attachments
    att = find_cfdi_attachments(uuid, comp.serie or "", comp.folio or "", comp.tipo_comprobante or "I")
    
    actual_path = comp.xml_path if comp and comp.xml_path and os.path.exists(comp.xml_path) else att["xml_path"]
              
    if not actual_path or not os.path.exists(actual_path):
        p_root = os.path.join(UPLOAD_DIR, f"{uuid}.xml")
        if os.path.exists(p_root): actual_path = p_root
        
        if not actual_path or not os.path.exists(actual_path):
             p_dup = os.path.join(LOGS_DIR, f"{uuid}.xml")
             if os.path.exists(p_dup): actual_path = p_dup

    if not actual_path or not os.path.exists(actual_path):
        raise HTTPException(status_code=404, detail="Archivo XML no encontrado")

    friendly_name = f"{(comp.serie or '') + (comp.folio or uuid)}.xml"
    return FileResponse(
        path=actual_path, 
        filename=friendly_name,
        media_type="text/xml",
        content_disposition_type="attachment",
        headers={"Cache-Control": "no-cache"}
    )

@router.get("/{uuid}/pdf")
async def get_comprobante_pdf(uuid: str, index: int = Query(0), db: AsyncSession = Depends(get_db)):
    if not uuid or len(uuid) < 32:
        raise HTTPException(status_code=400, detail="UUID inválido o malformado")

    try:
        query = select(Comprobante).where(func.lower(cast(Comprobante.uuid, String)) == uuid.lower())
        res = await db.execute(query)
        comp = res.scalars().first()
    except Exception as e_db:
        raise HTTPException(status_code=500, detail=f"Error de base de datos: {str(e_db)}")
    
    if not comp:
        raise HTTPException(status_code=404, detail="Comprobante no encontrado")

    actual_path = None
    pdf_raw = comp.pdf_path or ""
    pdf_list = [p for p in pdf_raw.split('|') if p.strip()]
    
    if index < len(pdf_list) and os.path.exists(pdf_list[index]):
        actual_path = pdf_list[index]
    else:
        for p in pdf_list:
            if os.path.exists(p):
                actual_path = p
                break
    
    if not actual_path:
        from src.services.cfdi_storage import get_cfdi_vault_path
        try:
            vault_dir = ""
            if comp.xml_path and os.path.exists(os.path.dirname(comp.xml_path)):
                vault_dir = os.path.dirname(comp.xml_path)
            else:
                vault_dir = get_cfdi_vault_path(str(comp.entidad_id), str(comp.rfc_emisor), comp.fecha_emision)
                
            if os.path.exists(vault_dir):
                for f in os.listdir(vault_dir):
                    if f.lower() == f"{str(comp.uuid).lower()}.pdf":
                        actual_path = os.path.join(vault_dir, f)
                        comp.pdf_path = actual_path
                        db.add(comp)
                        await db.commit()
                        break
        except Exception:
            pass

    if not actual_path or not os.path.exists(actual_path):
        if comp.folio == 'AR-7227':
            raise HTTPException(status_code=404, detail="ERROR: El binario del Pago AR-7227 no coincide con el UUID en disco.")
        
        attempted_dir = vault_dir if 'vault_dir' in locals() else "Desconocida/No generada"
        raise HTTPException(status_code=404, detail=f"ERROR: El binario del comprobante no coincide con el UUID en disco en {attempted_dir}")

    folio_raw = str(comp.folio or uuid)
    folio_clean = "".join([c for c in folio_raw if c.isdigit()]).lstrip('0') or '0'
    serie_clean = (comp.serie or "").replace("-", "")
    friendly_name = f"{serie_clean}{folio_clean}{f'_{index}' if index > 0 else ''}.pdf"
    
    return FileResponse(
        path=actual_path, 
        filename=friendly_name,
        media_type="application/pdf",
        content_disposition_type="attachment",
        headers={
            "Cache-Control": "no-store, no-cache, must-revalidate, max-age=0",
            "Pragma": "no-cache",
            "Expires": "0"
        }
    )


@router.post("/{uuid}/pdf")
async def upload_comprobante_pdf(
    uuid: str, 
    file: UploadFile = File(...), 
    db: AsyncSession = Depends(get_db)):
    
    if not uuid or len(uuid) < 32:
        raise HTTPException(status_code=400, detail="UUID inválido o malformado")
    
    try:
        query = select(Comprobante).where(func.lower(cast(Comprobante.uuid, String)) == uuid.lower())
        res = await db.execute(query)
        comp = res.scalars().first()
    except Exception as e_db:
        raise HTTPException(status_code=500, detail=f"Error de DB: {str(e_db)}")
    
    if not comp:
        raise HTTPException(status_code=404, detail="Comprobante no encontrado")
        
    vault_dir = ""
    if comp.xml_path and os.path.exists(os.path.dirname(comp.xml_path)):
        vault_dir = os.path.dirname(comp.xml_path)
    else:
        from src.services.cfdi_storage import get_cfdi_vault_path
        vault_dir = get_cfdi_vault_path(str(comp.entidad_id), str(comp.rfc_emisor), comp.fecha_emision)
        os.makedirs(vault_dir, exist_ok=True)
        
    target_path = os.path.join(vault_dir, f"{str(comp.uuid).lower()}.pdf")
    
    with open(target_path, "wb") as buffer:
        buffer.write(await file.read())
        
    comp.pdf_path = target_path
    db.add(comp)
    await db.commit()
    
    return {"status": "ok", "message": "Archivo inyectado físicamente en SSoT", "path": target_path}

@router.get("/{uuid}")
async def get_comprobante_detail(uuid: str, db: AsyncSession = Depends(get_db)):
    if not uuid or len(uuid) < 32:
        raise HTTPException(status_code=400, detail="UUID inválido o malformado")

    try:
        from src.database.models import CfdiRelacionado, CfdiConcepto
        from sqlalchemy.orm import selectinload
        query = select(Comprobante).options(selectinload(Comprobante.relacionados)).where(func.lower(cast(Comprobante.uuid, String)) == uuid.lower())
        res = await db.execute(query)
        comp = res.scalars().first()
        
        if not comp:
            raise HTTPException(status_code=404, detail="Comprobante no encontrado")

        total_final = float(comp.total or 0)
        if comp.tipo_comprobante == 'P':
            sum_q = select(func.sum(CfdiRelacionado.monto_pagado)).where(func.lower(CfdiRelacionado.uuid_padre) == uuid.lower())
            sum_res = await db.execute(sum_q)
            val_pago = sum_res.scalar()
            if val_pago is not None:
                 total_final = float(val_pago)

        conceptos_output = []
        
        try:
            import uuid as uuid_lib
            concept_q = select(CfdiConcepto.descripcion, CfdiConcepto.importe).where(
                CfdiConcepto.cfdi_id == uuid_lib.UUID(uuid)
            )
            concept_res = await db.execute(concept_q)
            for row in concept_res.all():
                conceptos_output.append({
                    "descripcion": row[0],
                    "importe": float(row[1] or 0)
                })
        except: pass

        if not conceptos_output:
            try:
                import xml.etree.ElementTree as ET
                from src.services.cfdi_storage import find_cfdi_attachments
                att = find_cfdi_attachments(uuid, comp.serie or "", comp.folio or "", comp.tipo_comprobante or "I")
                actual_path = comp.xml_path if comp and comp.xml_path and os.path.exists(comp.xml_path) else att["xml_path"]
                if actual_path:
                    with open(actual_path, "rb") as f_xml:
                        xml_content = f_xml.read().lstrip()
                    root = ET.fromstring(xml_content)
                    concept_nodes = root.findall('.//{*}Conceptos/{*}Concepto')
                    for c_node in concept_nodes:
                        desc = c_node.get('Descripcion') or c_node.get('descripcion') or 'Sin descripción'
                        imp = float(c_node.get('Importe') or c_node.get('importe') or 0.0)
                        conceptos_output.append({"descripcion": desc, "importe": imp})
            except: pass

        descripcion = " | ".join([c["descripcion"] for c in conceptos_output]) if conceptos_output else "Sin descripción"

        return {
            "uuid": str(comp.uuid),
            "serie": comp.serie or "",
            "folio": comp.folio or "",
            "tipo_comprobante": comp.tipo_comprobante or "I",
            "metodo_pago": comp.metodo_pago or "PUE",
            "forma_pago": comp.forma_pago or "03",
            "rfc_emisor": comp.rfc_emisor,
            "rfc_receptor": comp.rfc_receptor,
            "total": total_final,
            "descripcion_concepto": descripcion,
            "conceptos": conceptos_output,
            "pdf_count": len([p for p in (comp.pdf_path or "").split('|') if p.strip()]),
            "pdf_files": [os.path.basename(p) for p in (comp.pdf_path or "").split('|') if p.strip()]
        }
    except Exception as e:
        print(f"--- ERROR EN COMPROBANTE DETAIL --- \n{str(e)}")
        raise HTTPException(status_code=500, detail=f"Error en detalle del comprobante: {str(e)}")